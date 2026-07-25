#!/usr/bin/env python3
"""Sync Empleos_Tracker.xlsx <-> output/apply/apply-queue.csv

  python scripts/sync-empleos-tracker.py import   # Excel Pendiente (todos los canales) → cola
  python scripts/sync-empleos-tracker.py export   # cola → Excel Estado + Notas
"""

from __future__ import annotations

import csv
import re
import sys
import tempfile
import zipfile
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.styles import Font
except ImportError:
    print("Instalá openpyxl: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
QUEUE_PATH = ROOT / "output" / "apply" / "apply-queue.csv"
DEFAULT_XLSX = Path(r"C:\Users\gabri\OneDrive\Escritorio\Empleos_Tracker.xlsx")

JOB_ID_RE = re.compile(r"(?:currentJobId=|/jobs/view/)(\d+)")
ASSESSMENT_RE = re.compile(r"(assessment|honeypot)", re.I)

# Cola → Excel
# Automatización NUNCA escribe Descartado (solo la usuaria en Excel).
STATUS_TO_EXCEL = {
    "pendiente": "Pendiente",
    "enviada": "Enviada",
    "cerrada": "Cerrado",
    "descartada": "Stand-by",  # redirigido; ver nota abajo
    "duplicado": "Duplicado",
    "stand-by": "Stand-by",
    "borrador": "Borrador abierto",
}

# Excel → cola (solo lectura de finales / pendiente)
EXCEL_TO_STATUS = {
    "pendiente": "pendiente",
    "enviada": "enviada",
    "cerrado": "cerrada",
    "descartado": "descartada",
    "duplicado": "duplicado",
    "stand-by": "pendiente",
    "borrador abierto": "pendiente",
}


def job_id_from_url(url: str) -> str:
    m = JOB_ID_RE.search(url or "")
    return m.group(1) if m else ""


def match_percent(raw) -> int:
    if raw is None:
        return 0
    s = str(raw).strip().replace("%", "")
    try:
        return int(float(s))
    except ValueError:
        return 0


def header_map(ws) -> dict[str, int]:
    """Nombre de columna (lower) → índice 1-based."""
    out: dict[str, int] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        v = ws.cell(1, col).value
        if v is None:
            continue
        out[str(v).strip().lower()] = col
    return out


def ensure_notas_column(ws) -> int:
    headers = header_map(ws)
    if "notas" in headers:
        return headers["notas"]
    col = (ws.max_column or 0) + 1
    ws.cell(1, col).value = "Notas"
    return col


def ensure_mis_comentarios_column(ws) -> int:
    """Columna solo-usuaria (#168). Crea header si falta; nunca escribe celdas de datos."""
    headers = header_map(ws)
    if "mis comentarios" in headers:
        return headers["mis comentarios"]
    # Preferir col 11 si está libre / es la siguiente tras Notas
    notas_col = headers.get("notas") or 10
    col = max(notas_col + 1, 11)
    # Si otra columna ya usa ese índice con otro nombre, append al final
    if col in headers.values() and "mis comentarios" not in headers:
        col = (ws.max_column or 0) + 1
    ws.cell(1, col).value = "Mis comentarios"
    return col


def notes_cell_value(text: str):
    """Si menciona assessment/honeypot, dejar esa palabra en negrita (rich text)."""
    if not text:
        return text
    if not ASSESSMENT_RE.search(text):
        return text
    parts: list[TextBlock | str] = []
    pos = 0
    for m in ASSESSMENT_RE.finditer(text):
        if m.start() > pos:
            parts.append(text[pos : m.start()])
        parts.append(TextBlock(InlineFont(b=True), m.group(0)))
        pos = m.end()
    if pos < len(text):
        parts.append(text[pos:])
    try:
        return CellRichText(*parts)
    except Exception:
        # Fallback: celda entera en negrita
        return text


def cell_as_text(value) -> str:
    """Normaliza celda (str / rich text) a texto plano."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    try:
        return str(value).strip()
    except Exception:
        return ""


def append_notes(existing, incoming: str) -> str:
    """Append Notas sin pisar ni duplicar (B30 / sync tracker)."""
    prev = cell_as_text(existing)
    new = (incoming or "").strip()
    if not new:
        return prev
    if not prev:
        return new
    if new in prev:
        return prev
    return f"{prev}\n{new}"


# Estados que la automatización no pisa (usuaria / finales / stand-by).
PROTECTED_ESTADOS = frozenset(
    {
        "cerrado",
        "descartado",
        "stand-by",
        "duplicado",
        "enviada",
        "en proceso",
        "seleccionada",
        "a-pendiente",
        "a-realizado",
    }
)


def fix_absolute_table_targets(xlsx_path: Path) -> bool:
    """openpyxl → Target='/xl/tables/…'; ExcelJS necesita '../tables/…'."""
    tmp = Path(tempfile.mkdtemp())
    with zipfile.ZipFile(xlsx_path, "r") as z:
        z.extractall(tmp)
    changed = False
    sheets = tmp / "xl" / "worksheets"
    if sheets.exists():
        for rels in sheets.rglob("*.rels"):
            txt = rels.read_text(encoding="utf-8")
            new = re.sub(
                r'Target="/xl/tables/(table\d+\.xml)"',
                r'Target="../tables/\1"',
                txt,
            )
            if new != txt:
                rels.write_text(new, encoding="utf-8")
                changed = True
    if not changed:
        return False
    out = xlsx_path.with_suffix(".xlsx.tmp")
    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for p in tmp.rglob("*"):
            if p.is_file():
                z.write(p, p.relative_to(tmp).as_posix())
    out.replace(xlsx_path)
    return True


def easy_apply_from_canal(canal: str) -> str:
    c = (canal or "").strip().lower()
    if c == "easy apply":
        return "yes"
    if c == "externo":
        return "no"
    return ""


def cmd_import(xlsx: Path) -> None:
    existing = load_queue()
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Empleos"]
    pending_rows: list[dict] = []
    ea_n = ext_n = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[3]:
            continue
        estado = (r[5] or "").strip()
        if estado.lower() != "pendiente":
            continue
        url = str(r[3]).strip()
        jid = job_id_from_url(url)
        if not jid:
            continue
        canal = (r[4] or "").strip()
        ea = easy_apply_from_canal(canal)
        if ea == "yes":
            ea_n += 1
        elif ea == "no":
            ext_n += 1
        pending_rows.append(
            {
                "jobId": jid,
                "matchPercent": match_percent(r[0]),
                "title": (r[1] or "").strip(),
                "company": (r[2] or "").strip(),
                "url": f"https://www.linkedin.com/jobs/view/{jid}/",
                "easyApply": ea,
                "status": "pendiente",
                "reason": "Importado desde Empleos_Tracker.xlsx",
                "notes": "",
                "updatedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            }
        )
    pending_rows.sort(key=lambda x: -x["matchPercent"])

    # Conservar filas finales de la cola; refrescar todos los Pendiente desde Excel.
    by_id: dict[str, dict] = {}
    for jid, row in existing.items():
        st = (row.get("ApplyStatus") or "pendiente").strip().lower()
        if st in ("pendiente", "pending", ""):
            continue
        by_id[jid] = row

    for p in pending_rows:
        jid = p["jobId"]
        prev = existing.get(jid, {})
        notes = (prev.get("Notes") or prev.get("notes") or "").strip()
        by_id[jid] = {
            "JobId": jid,
            "Match%": str(p["matchPercent"]),
            "Title": p["title"],
            "Company": p["company"],
            "URL": p["url"],
            "EasyApply": p["easyApply"],
            "ApplyStatus": "pendiente",
            "Reason": p["reason"],
            "Notes": notes,
            "UpdatedAt": p["updatedAt"],
        }

    merged = sorted(
        by_id.values(),
        key=lambda row: -match_percent(row.get("Match%")),
    )

    QUEUE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with QUEUE_PATH.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(
            [
                "JobId",
                "Match%",
                "Title",
                "Company",
                "URL",
                "EasyApply",
                "ApplyStatus",
                "Reason",
                "Notes",
                "UpdatedAt",
            ]
        )
        for row in merged:
            w.writerow(
                [
                    row.get("JobId", ""),
                    row.get("Match%", ""),
                    row.get("Title", ""),
                    row.get("Company", ""),
                    row.get("URL", ""),
                    row.get("EasyApply", ""),
                    row.get("ApplyStatus", ""),
                    row.get("Reason", ""),
                    row.get("Notes", ""),
                    row.get("UpdatedAt", ""),
                ]
            )
    print(
        f"Import OK: {len(pending_rows)} Pendiente desde Excel "
        f"({ea_n} Easy Apply, {ext_n} externo) · cola total {len(merged)} → {QUEUE_PATH}"
    )


def load_queue() -> dict[str, dict]:
    if not QUEUE_PATH.exists():
        return {}
    by_id = {}
    with QUEUE_PATH.open(encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            jid = (row.get("JobId") or "").strip()
            if jid:
                by_id[jid] = row
    return by_id


def cmd_export(xlsx: Path) -> None:
    queue = load_queue()
    if not queue:
        print("Cola vacía; nada que exportar.")
        return

    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Empleos"]
    notas_col = ensure_notas_column(ws)
    ensure_mis_comentarios_column(ws)  # schema only; never write data cells
    headers = header_map(ws)
    proximo_col = headers.get("próximo paso") or headers.get("proximo paso")

    updated = 0
    notes_updated = 0
    skipped_final = 0
    for row_idx in range(2, ws.max_row + 1):
        url = ws.cell(row_idx, 4).value  # LinkedIn
        jid = job_id_from_url(str(url or ""))
        if not jid or jid not in queue:
            continue
        current = (ws.cell(row_idx, 6).value or "").strip()  # Estado
        current_l = current.lower()
        q = queue[jid]
        status = (q.get("ApplyStatus") or "pendiente").strip().lower()
        excel_estado = STATUS_TO_EXCEL.get(status)
        # Defensa: nunca escribir Descartado desde la cola
        if excel_estado and excel_estado.lower() == "descartado":
            excel_estado = "Stand-by"
            reason_extra = "Automatización no puede Descartado → Stand-by"
            notes_q = (q.get("Notes") or q.get("notes") or "").strip()
            if reason_extra not in notes_q:
                q["Notes"] = f"{notes_q}\n{reason_extra}".strip() if notes_q else reason_extra
        if status == "descartada":
            notes_q = (q.get("Notes") or q.get("notes") or "").strip()
            hint = "Stand-by: cola tenía descartada (solo vos marcás Descartado en Excel)"
            if hint not in notes_q:
                q["Notes"] = f"{notes_q}\n{hint}".strip() if notes_q else hint

        # No pisar estados protegidos (Enviada / finales / stand-by / usuaria)
        status_locked = current_l in PROTECTED_ESTADOS
        if status_locked:
            skipped_final += 1
        elif excel_estado and current != excel_estado:
            ws.cell(row_idx, 6).value = excel_estado
            if excel_estado == "Enviada" and not ws.cell(row_idx, 7).value:
                ws.cell(row_idx, 7).value = datetime.now().strftime("%Y-%m-%d")
            updated += 1

        notes = (q.get("Notes") or q.get("notes") or "").strip()
        reason = (q.get("Reason") or "").strip()
        # Si no hay Notes pero Reason habla de campos/preguntas/assessment, volcar Reason
        if not notes and re.search(
            r"preguntas nuevas|assessment|honeypot|definir respuesta|"
            r"a[nñ]os de experiencia|deequ|great expectations|"
            r"campos (que )?fallaron|campos sin completar|typeahead|dropdown|"
            r"skill no mapeada|pendiente",
            reason,
            re.I,
        ):
            notes = reason

        if notes:
            cell = ws.cell(row_idx, notas_col)
            merged = append_notes(cell.value, notes)
            if merged != cell_as_text(cell.value):
                cell.value = notes_cell_value(merged)
                if ASSESSMENT_RE.search(merged) and not isinstance(cell.value, CellRichText):
                    cell.font = Font(bold=True)
                notes_updated += 1
            if proximo_col and "Preguntas nuevas" in notes:
                prev = (ws.cell(row_idx, proximo_col).value or "").strip()
                hint = "Definir respuestas (ver Notas) y avisar en chat"
                if hint.lower() not in prev.lower():
                    ws.cell(row_idx, proximo_col).value = (
                        f"{prev} | {hint}" if prev else hint
                    )

    wb.save(xlsx)
    if fix_absolute_table_targets(xlsx):
        print("   (Target tabla normalizado para ExcelJS)")
    print(
        f"Export OK: {updated} estados, {notes_updated} notas en {xlsx} "
        f"(estados protegidos omitidos: {skipped_final})"
    )


def main() -> None:
    if len(sys.argv) < 2 or sys.argv[1] not in ("import", "export"):
        print(__doc__)
        sys.exit(1)
    xlsx = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_XLSX
    if not xlsx.exists():
        print(f"No existe: {xlsx}")
        sys.exit(1)
    if sys.argv[1] == "import":
        cmd_import(xlsx)
    else:
        cmd_export(xlsx)


if __name__ == "__main__":
    main()
